from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login as auth_login, logout as auth_logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, F, DecimalField, OuterRef, Subquery, Value
from django.db.models.functions import Coalesce
from django.db import transaction
from django.utils import timezone
from django_tables2 import RequestConfig
from django.forms import inlineformset_factory
from django.http import HttpResponse
from quermesse import tables
from quermesse import models
from quermesse import forms
import openpyxl
import os
from decimal import Decimal
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

ItemCaixaCreatFormSet = inlineformset_factory(
    models.Caixa,
    models.ItemCaixa,
    fields=('produtos','quantidade'),
    extra=20,
    can_delete=True
)

ItemCaixaEditFormSet = inlineformset_factory(
    models.Caixa,
    models.ItemCaixa,
    fields=('produtos','quantidade'),
    extra=0,
    can_delete=True
)

ItemCortesiaCreatFormSet = inlineformset_factory(
    models.Cortesia,
    models.ItemCortesia,
    fields=('produtos','quantidade'),
    extra=20,
    can_delete=True
)

ItemCortesiaEditFormSet = inlineformset_factory(
    models.Cortesia,
    models.ItemCortesia,
    fields=('produtos','quantidade'),
    extra=0,
    can_delete=True
)

def login(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            auth_login(request, user)
            return redirect('home')
    else:
        form = AuthenticationForm()
    return render(request, 'login/login.html', {'form': form})

def logout(request):
    auth_logout(request)
    return redirect('login')

@login_required
def gerador_bingo(request):
    api_key = os.getenv('API_KEY', default='')
    return render(request, 'gerador_bingo/index.html', {
        'api_key': api_key
    })

@login_required
def home(request):
    qs_produtos = (
        models.ItemCaixa.objects
        .values(produto=F('produtos__nome'))
        .annotate(
            total_qtd = Sum('quantidade'),
            total_valor = Sum(
                F('quantidade') * F('produtos__valor'),
                output_field=DecimalField(decimal_places=2)
            )
        )
        .order_by('-total_qtd')
    )
    qs_operador = (
        models.Clientes.objects
                .filter(is_caixa=True)
                .annotate(total_vendas=Sum("caixa__valor"))
                .order_by("-total_vendas")
    )
    qs_fiado_pago = models.Fiado.objects.filter(is_pago=True)
    qs_fiado_aberto = models.Fiado.objects.filter(is_pago=False)
    qs_caixa = models.Caixa.objects.all()
    qs_entrada = models.Entradas.objects.all()
    qs_despesa = models.Despesas.objects.all()
    soma_fiado_pago = qs_fiado_pago.aggregate(total_valor_pago=Sum('valor'))['total_valor_pago'] or Decimal('0.00')
    soma_fiado_aberto = qs_fiado_aberto.aggregate(total_valor_aberto=Sum('valor'))['total_valor_aberto'] or Decimal('0.00')
    soma_total_caixa = qs_caixa.aggregate(total_valor_caixa=Sum('valor'))['total_valor_caixa'] or Decimal('0.00')
    soma_total_entrada = qs_entrada.aggregate(total_valor_entrada=Sum('valor'))['total_valor_entrada'] or Decimal('0.00')
    soma_total_despesa = qs_despesa.aggregate(total_valor_despesa=Sum('valor'))['total_valor_despesa'] or Decimal('0.00')
    soma_total_bruto = soma_total_caixa + soma_total_entrada
    soma_total_liquido = soma_total_bruto - soma_total_despesa
    soma_total_fiado = soma_fiado_aberto + soma_fiado_pago
    table_produto = tables.QtdProdutosTable(qs_produtos)
    table_operador = tables.OperadorTotalTable(qs_operador)
    return render(request, 'dashboard/dashboard.html', {
        'title': 'Dashboard',
        'soma_fiado_pago': soma_fiado_pago,
        'soma_fiado_aberto': soma_fiado_aberto,
        'soma_total_fiado': soma_total_fiado,
        'soma_total_caixa': soma_total_caixa,
        'soma_total_entrada': soma_total_entrada,
        'soma_total_despesa': soma_total_despesa,
        'soma_total_bruto': soma_total_bruto,
        'soma_total_liquido': soma_total_liquido,
        'table_produto': table_produto,
        'table_operador': table_operador
    })

@login_required
def clientes(request):
    clientes = models.Clientes.objects.filter(is_cliente=True).order_by('nome').all()
    table = tables.ClientesTable(clientes)
    RequestConfig(request, paginate={"per_page": 15}).configure(table)
    return render(request, 'clientes/clientes.html', {
        'title': 'Clientes',
        'table': table
    })

@login_required
def add_cliente(request):
    if request.method == 'POST':
        form = forms.ClientesForm(request, request.POST)
        if form.is_valid():
            novo_cliente = form.save(commit=False)
            novo_cliente.is_cliente = form.cleaned_data.get('is_cliente', True)
            novo_cliente.is_caixa = form.cleaned_data.get('is_caixa', False)
            novo_cliente.create_user = form.cleaned_data.get('create_user', request.user)
            novo_cliente.assign_user = form.cleaned_data.get('assign_user', request.user)
            novo_cliente.save()
            messages.success(request, f'O cliente {novo_cliente.nome} foi adicionado com sucesso!')
            return redirect('clientes')
    else:
        form = forms.ClientesForm(request)
    return render(request, 'clientes/add_cliente.html', {
        'title': 'Adicionar Cliente',
        'form': form
    })

@login_required
def edit_cliente(request, cliente_id):
    qs_cliente = get_object_or_404(models.Clientes, pk=cliente_id)
    if request.method == 'POST':
        form = forms.ClientesEditForm(request.POST, instance=qs_cliente)
        if form.is_valid():
            cliente = form.save(commit=False)
            cliente.assign_user = form.cleaned_data.get('assign_user', request.user)
            cliente.save()
            messages.success(request, f'O cliente {cliente.nome} foi editado com sucesso!')
            return redirect('clientes')
    else:
        form = forms.ClientesEditForm(instance=qs_cliente)
    return render(request, 'clientes/add_cliente.html', {
        'title': 'Editar cliente',
        'form': form
    })

@login_required
def delete_cliente_modal(request):
    pk = request.GET.get('id')
    record = get_object_or_404(models.Clientes, id=pk)
    return render(request, 'clientes/confirmar_deletar.html', {
        'record': record
    })

@login_required
def delete_cliente(request, cliente_id):
    cliente = get_object_or_404(models.Clientes, pk=cliente_id)
    if request.method == 'POST':
        cliente.delete()
        messages.success(request, 'O cliente foi excluido com sucesso!')
        return redirect('clientes')
    return HttpResponse(status=405)

@login_required
def autorizados(request):
    autorizados = models.ClienteUsuario.objects.order_by('cliente__nome').all()
    table = tables.AutorizadoTable(autorizados)
    RequestConfig(request, paginate={"per_page": 15}).configure(table)
    return render(request, 'autorizados/autorizados.html', {
        'title': 'Autorizados',
        'table': table
    })

@login_required
def add_autorizado(request):
    if request.method == 'POST':
        form = forms.AutorizadoForm(request.POST)
        if form.is_valid():
            novo_autorizado = form.save(commit=False)
            novo_autorizado.create_user = form.cleaned_data.get('create_user', request.user)
            novo_autorizado.assign_user = form.cleaned_data.get('assign_user', request.user)
            novo_autorizado.save()
            messages.success(request, f'O novo autorizado: "{novo_autorizado}" adicionado com sucesso!')
            return redirect('autorizados')
    else:
        form = forms.AutorizadoForm()
    return render(request, 'autorizados/add_autorizados.html', {
        'title': 'Adicionar autorizados',
        'form': form
    })

@login_required
def edit_autorizado(request, autorizado_id):
    qs_autorizado = get_object_or_404(models.ClienteUsuario, pk=autorizado_id)
    if request.method == 'POST':
        form = forms.AutorizadoForm(request.POST, instance=qs_autorizado)
        if form.is_valid():
            autorizado = form.save(commit=False)
            autorizado.assign_user = form.cleaned_data.get('assign_user', request.user)
            messages.success(request, f'O usuario autorizado: "{autorizado}" foi editado com sucesso!')
            return redirect('autorizados')
    else:
        form = forms.AutorizadoForm(instance=qs_autorizado)
    return render(request, 'autorizados/add_autorizados.html', {
        'title': 'Editar autorizado',
        'form': form
    })

@login_required
def delete_autorizado_modal(request):
    pk = request.GET.get('id')
    record = get_object_or_404(models.ClienteUsuario, id=pk)
    return render(request, 'autorizados/confirmar_deletar.html', {
        'record': record
    })

@login_required
def delete_autorizado(request, autorizado_id):
    autorizado = get_object_or_404(models.ClienteUsuario, pk=autorizado_id)
    if request.method == 'POST':
        autorizado.delete()
        messages.success(request, 'O usuario autorizado foi deletado com sucesso!')
        return redirect('autorizados')
    return HttpResponse(status=405)

@login_required
def fiados(request):
    if request.method == 'POST':
        selecionados = request.POST.getlist('select')
        if selecionados:
            models.Fiado.objects.filter(pk__in=selecionados).update(
                is_pago=True,
                datapago=timezone.now()
            )
            messages.success(request, f"{len(selecionados)} fiado(s) marcado(s) como pago.")
        else:
            messages.warning(request, "Nenhum registro selecionado.")
        return redirect('fiados')
    form = forms.FindFiadoForm(request.GET)
    form.fields['cliente'].required = False
    form.fields['datadoc'].required = False
    filter_search = {}
    if form.is_valid():
        cliente = form.cleaned_data.get('cliente')
        datadoc = form.cleaned_data.get('datadoc')
        datapago = form.cleaned_data.get('datapago')
        is_pago = form.cleaned_data.get('is_pago')
        if cliente:
            filter_search['cliente'] = cliente
        if datadoc:
            filter_search['datadoc'] = datadoc
        if datapago:
            filter_search['datapago'] = datapago
        if is_pago is not None:
            filter_search['is_pago'] = is_pago
    fiados = models.Fiado.objects.filter(**filter_search).order_by('cliente__nome')

    fiados_agrupados = (
        fiados.values('cliente__nome')
        .annotate(total_fiado=Sum('valor'))
        .order_by('cliente__nome')
    )

    if request.GET.get('export') == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Fiados por cliente'

        ws.append(['Nome do cliente', "Total"])

        for row in fiados_agrupados:
            ws.append([
                row['cliente__nome'],
                float(row['total_fiado'] or 0)
            ])

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

        row_count = len(fiados_agrupados) + 1
        header_font = Font(bold=True, color='FFFFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='4F81BD')

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, max_row=row_count):
            for cell in row:
                cell.number_format = '"R$"#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="fiados.xlsx"'
        wb.save(response)
        return response
    soma_valor = fiados.aggregate(total_valor=Sum('valor'))['total_valor'] or Decimal('0.00')
    table = tables.FiadosTable(fiados)
    RequestConfig(request, paginate={"per_page": 15}).configure(table)
    return render(request, 'fiados/fiados.html', {
        'title': 'Fiados',
        'soma_valor': soma_valor,
        'table': table,
        'form': form
    })

@login_required
def add_fiado(request):
    if request.method == 'POST':
        form = forms.FiadoForm(request.POST)
        if form.is_valid():
            novo_fiado = form.save(commit=False)
            novo_fiado.is_pago = False
            novo_fiado.creat_user = form.cleaned_data.get('create_user', request.user)
            novo_fiado.assign_user = form.cleaned_data.get('assign_user', request.user)
            novo_fiado.save()
            messages.success(request, f'O fiado do cliente {novo_fiado.cliente} foi adicionado com sucesso!')
            return redirect('fiados')
    else:
        form = forms.FiadoForm()
    return render(request, 'fiados/add_fiado.html', {
        'title': 'Adicionar fiado',
        'form': form
    })

@login_required
def edit_fiado(request, fiado_id):
    qs_fiado = get_object_or_404(models.Fiado, pk=fiado_id)
    if request.method == 'POST':
        form = forms.FiadoEditForm(request.POST, instance=qs_fiado)
        if form.is_valid():
            fiado = form.save(commit=False)
            fiado.assign_user = form.cleaned_data.get('assign_user', request.user)
            fiado.save()
            messages.success(request, f'O fiado do cliente {fiado.cliente} foi editado com sucesso!')
            return redirect('fiados')
    else:
        form = forms.FiadoEditForm(instance=qs_fiado)
    return render(request, 'fiados/add_fiado.html', {
        'title': 'Editar fiado',
        'form': form
    })

@login_required
def delete_fiado_modal(request):
    pk = request.GET.get('id')
    record = get_object_or_404(models.Fiado, id=pk)
    return render(request, 'fiados/confirmar_deletar.html', {
        'record': record
    })

@login_required
def delete_fiado(request, fiado_id):
    fiado = get_object_or_404(models.Fiado, pk=fiado_id)
    if request.method == 'POST':
        fiado.delete()
        messages.success(request, 'O fiado foi excluido com sucesso!')
        return redirect('fiados')
    return HttpResponse(status=405)

@login_required
def operadores(request):
    operadores = models.Clientes.objects.filter(is_caixa=True).order_by('nome').all()
    table = tables.OperadorTable(operadores)
    RequestConfig(request, paginate={"per_page": 15}).configure(table)
    return render(request, 'operadores/operadores.html', {
        'title': 'Operadores',
        'table': table
    })

@login_required
def add_operador(request):
    if request.method == 'POST':
        form = forms.OperadoresForm(request, request.POST)
        if form.is_valid():
            novo_operador = form.save(commit=False)
            novo_operador.is_caixa = form.cleaned_data.get('is_caixa', True)
            novo_operador.is_cliente = form.cleaned_data.get('is_cliente', False)
            novo_operador.create_user = form.cleaned_data.get('create_user', request.user)
            novo_operador.assign_user = form.cleaned_data.get('assign_user', request.user)
            novo_operador.save()
            messages.success(request, f'O operador {novo_operador.nome} foi adicionado com sucesso!')
            return redirect('operadores')
    else:
        form = forms.OperadoresForm(request)
    return render(request, 'operadores/add_operador.html', {
        'title': 'Adicionar operador',
        'form': form
    })

@login_required
def edit_operador(request, operador_id):
    qs_operador = get_object_or_404(models.Clientes, pk=operador_id)
    if request.method == 'POST':
        form = forms.OperadoresEditForm(request.POST, instance=qs_operador)
        if form.is_valid():
            operador = form.save(commit=False)
            operador.assign_user = form.cleaned_data.get('assign_user', request.user)
            operador.save()
            messages.success(request, f'O operador {operador.nome} foi editado com sucesso!')
            return redirect('operadores')
    else:
        form = forms.OperadoresEditForm(instance=qs_operador)
    return render(request, 'operadores/add_operador.html', {
        'title': 'Editar operador',
        'form': form
    })

@login_required
def delete_operador_modal(request):
    pk = request.GET.get('id')
    record = get_object_or_404(models.Clientes, id=pk)
    return render(request, 'operadores/confirmar_deletar.html', {
        'record': record
    })

@login_required
def delete_operador(request, operador_id):
    operador = get_object_or_404(models.Clientes, pk=operador_id)
    if request.method == 'POST':
        operador.delete()
        messages.success(request, 'O operador foi deletado com sucesso!')
        return redirect('operadores')
    return HttpResponse(status=405)

@login_required
def produtos(request):
    produtos = models.Produto.objects.order_by('nome').all()
    table = tables.ProdutosTable(produtos)
    RequestConfig(request, paginate={"per_page": 15}).configure(table)
    return render(request, 'produtos/produtos.html', {
        'title': 'Produtos',
        'table': table
    })

@login_required
def add_produto(request):
    if request.method == 'POST':
        form = forms.ProdutoForm(request.POST)
        if form.is_valid():
            novo_produto = form.save(commit=False)
            novo_produto.create_user = form.cleaned_data.get('create_user', request.user)
            novo_produto.assign_user = form.cleaned_data.get('assign_user', request.user)
            novo_produto.save()
            messages.success(request, f'O produto {novo_produto.nome} foi adicionado com sucesso!')
            return redirect('produtos')
    else:
        form = forms.ProdutoForm()
    return render(request, 'produtos/add_produto.html', {
        'title': 'Adicionar produto',
        'form': form
    })

@login_required
def edit_produto(request, produto_id):
    qs_produto = get_object_or_404(models.Produto, pk=produto_id)
    if request.method == 'POST':
        form = forms.ProdutoForm(request.POST, instance=qs_produto)
        if form.is_valid():
            produto = form.save(commit=False)
            produto.assign_user = form.cleaned_data.get('assign_user', request.user)
            produto.save()
            messages.success(request, f'O produto {produto.nome} foi editado com sucesso!')
            return redirect('produtos')
    else:
        form = forms.ProdutoForm(instance=qs_produto)
    return render(request, 'produtos/add_produto.html', {
        'title': 'Editar produto',
        'form': form
    })

@login_required
def delete_produto_modal(request):
    pk = request.GET.get('id')
    record = get_object_or_404(models.Produto, id=pk)
    return render(request, 'produtos/confirmar_deletar.html', {
        'record': record
    })

@login_required
def delete_produto(request, produto_id):
    produto = get_object_or_404(models.Produto, pk=produto_id)
    if request.method == 'POST':
        produto.delete()
        messages.success(request, 'O produto foi deletado com sucesso!')
        return redirect('produtos')
    return HttpResponse(status=405)

@login_required
def total_produtos(request):
    cortesia_qs = (
        models.ItemCortesia.objects
        .filter(produtos=OuterRef('produtos'))
        .values('produtos')
        .annotate(
            cortesias_qtd = Sum('quantidade'),
            cortesias_valor = Sum(
                F('quantidade') * F('produtos__valor'),
                output_field=DecimalField(decimal_places=2)
            )
        )
        .values('cortesias_qtd', 'cortesias_valor')
    )
    qs = (
        models.ItemCaixa.objects
        .values('produtos', produto=F('produtos__nome'))
        .annotate(
            total_qtd = Sum('quantidade'),
            total_valor = Sum(
                F('quantidade') * F('produtos__valor'),
                output_field=DecimalField(decimal_places=2)
            ),
            cortesias_qtd = Coalesce(Subquery(cortesia_qs.values('cortesias_qtd')), Value(0)),
            cortesias_valor = Coalesce(Subquery(cortesia_qs.values('cortesias_valor')), 
                                        Value(0, output_field=DecimalField(decimal_places=2)),
                                        output_field=DecimalField(decimal_places=2)
                                        )
        )
        .annotate(
            qtd_liquida = F('total_qtd') - F('cortesias_qtd'),
            valor_liquido = F('total_valor') - F('cortesias_valor'),
        )
        .order_by('produto')
    )
    if request.GET.get('export') == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Produtos'
        ws.append(
            ['Nome produto', 
            'Quantidade vendida',
            'Quantidade de cortesias',
            'Quantidade líquida',
            'Valor total',
            'Valor total cortesias',
            'Valor total líquido',
            ])

        for row in qs:
            ws.append([
                row['produto'],
                row['total_qtd'],
                row['cortesias_qtd'],
                row['qtd_liquida'],
                row['total_valor'],
                row['cortesias_valor'],
                float(row['valor_liquido'] or 0)
            ])

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 30
        ws.freeze_panes = 'A2'

        header_font = Font(bold=True, color='FFFFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='4F81BD')

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row_count = len(qs) + 1
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=7, max_row=row_count):
            for cell in row:
                cell.number_format = '"R$"#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="produtos.xlsx"'
        wb.save(response)
        return response
    table = tables.QtdProdutosTable(qs)
    RequestConfig(request, paginate={"per_page": 25}).configure(table)
    return render(request, 'produtos/total_produtos.html', {
        'title': 'Vendas de produtos',
        'table': table
    })

@login_required
def caixas(request):
    form = forms.CaixaFindForm(request.GET)
    form.fields['cliente'].required = False
    form.fields['data'].required = False
    filter_search = {}
    if form.is_valid():
        cliente = form.cleaned_data.get('cliente')
        data = form.cleaned_data.get('data')
        if cliente:
            filter_search['cliente'] = cliente
        if data:
            filter_search['data'] = data
    caixas = models.Caixa.objects.filter(**filter_search).order_by('cliente__nome').all()
    caixas_fiados = models.Caixa.objects.filter(**filter_search).filter(cliente__nome__icontains='fiado')
    caixas_agrupados = (
        caixas.values('cliente__id', 'cliente__nome')
        .annotate(
            total_caixa = Sum('valor'),
            total_dinheiro = Sum('qtd_dinheiro'),
            total_cd = Sum('qtd_cd'),
            total_cc = Sum('qtd_cc'),
            total_pix = Sum('pix'),
            total_reimpressao = Sum('reimpressao')
        )
        .order_by('cliente__nome')
    )
    if request.GET.get('export') == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Caixas'
        ws.append(['Nome operador', 'Total', 
                    'Total em dinheiro', 'Total em catão de débito',
                    'Total em catão de crédito', 'Total em pix',
                    'Total de reimpressão'
                ])
        for row in caixas_agrupados:
            ws.append([
                row['cliente__nome'],
                row['total_caixa'],
                row['total_dinheiro'],
                row['total_cd'],
                row['total_cc'],
                row['total_pix'],
                row['total_reimpressao']
            ])

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 30
        ws.freeze_panes = 'A2'

        row_count = len(caixas_agrupados) + 1
        header_font = Font(bold=True, color='FFFFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='4F81BD')

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in ws.iter_rows(min_row=2, min_col=2, max_col=7, max_row=row_count):
            for cell in row:
                cell.number_format = '"R$"#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')

        table_ref = f"A1:G{row_count}"
        tab = Table(displayName="Caixas", ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

        total_row = row_count + 2
        cell_head = ws.cell(row=total_row, column=1, value="Total Geral:")
        cell_head.font = Font(bold=True)
        cell_head = Alignment(horizontal="right")
        cell = ws.cell(
            row=total_row,
            column=2,
            value="=SUM(Caixas[Total])"
        )
        cell.number_format = '"R$"#,##0.00'
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right")

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="caixas.xlsx"'
        wb.save(response)
        return response

    soma_valor = caixas.aggregate(total_valor=Sum('valor'))['total_valor'] or Decimal('0.00')
    soma_dinheiro_interno = caixas.aggregate(total_valor=Sum('qtd_dinheiro'))['total_valor'] or Decimal('0.00')
    soma_reimpressao_interno = caixas.aggregate(total_valor_reimpressao=Sum('reimpressao'))['total_valor_reimpressao'] or Decimal('0.00')
    soma_dinheiro = soma_dinheiro_interno - soma_reimpressao_interno
    soma_cd = caixas.aggregate(total_valor=Sum('qtd_cd'))['total_valor'] or Decimal('0.00')
    soma_cc = caixas.aggregate(total_valor=Sum('qtd_cc'))['total_valor'] or Decimal('0.00')
    soma_pix = caixas.aggregate(total_valor=Sum('pix'))['total_valor'] or Decimal('0.00')
    soma_reimpressao = caixas.aggregate(total_valor_reimpressao=Sum('reimpressao'))['total_valor_reimpressao'] or Decimal('0.00')
    soma_caixa_fiado = caixas_fiados.aggregate(total_fiado=Sum('valor'))['total_fiado'] or Decimal('0.00')
    table = tables.CaixaTable(caixas)
    RequestConfig(request, paginate={"per_page": 25}).configure(table)
    return render(request, 'caixas/caixas.html', {
        'title': 'Caixas',
        'form': form,
        'table': table,
        'soma_valor': soma_valor,
        'soma_dinheiro': soma_dinheiro,
        'soma_cd': soma_cd,
        'soma_cc': soma_cc,
        'soma_pix': soma_pix,
        'soma_reimpressao': soma_reimpressao,
        'soma_caixa_fiado': soma_caixa_fiado
    })

@login_required
def add_caixa(request):
    if request.method == 'POST':
        form = forms.CaixaForm(request.POST)
        formset = ItemCaixaCreatFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            caixa = form.save(commit=False)
            caixa.create_user = form.cleaned_data.get('create_user', request.user)
            caixa.assign_user = form.cleaned_data.get('assign_user', request.user)
            caixa.save()
            formset.instance = caixa
            formset.save()
            messages.success(request, 'O novo registro do caixa foi feito com sucesso!')
            return redirect('caixas')
    else:
        form = forms.CaixaForm()
        formset = ItemCaixaCreatFormSet()
    return render(request, 'caixas/add_caixa.html', {
        'title': 'Adicionar novo registro',
        'form': form,
        'formset': formset
    })

@login_required
def edit_caixa(request, caixa_id):
    caixa = get_object_or_404(models.Caixa, pk=caixa_id)
    if request.method == 'POST':
        form = forms.CaixaForm(request.POST, instance=caixa)
        formset = ItemCaixaEditFormSet(request.POST, instance=caixa)
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                caixa = form.save(commit=False)
                caixa.assign_user = request.user
                caixa.save()
                formset.save()
            messages.success(request, 'O registro do caixa foi editado com sucesso!')
            return redirect('caixas')
    else:
        form    = forms.CaixaForm(instance=caixa)
        formset = ItemCaixaEditFormSet(instance=caixa)
    return render(request, 'caixas/add_caixa.html', {
        'title':   'Editar registro do caixa',
        'form':    form,
        'formset': formset,
    })

@login_required
def delete_caixa_modal(request):
    pk = request.GET.get('id')
    record = get_object_or_404(models.Caixa, id=pk)
    return render(request, 'caixas/confirmar_deletar.html', {
        'record': record
    })

@login_required
def delete_caixa(request, caixa_id):
    qs_caixa = get_object_or_404(models.Caixa, pk=caixa_id)
    if request.method == 'POST':
        qs_caixa.delete()
        messages.success(request, 'O registro do caixa foi deletado com sucesso!')
        return redirect('caixas')
    return HttpResponse(status=405)

@login_required
def cortesia(request):
    cortesias = models.Cortesia.objects.order_by('-data').all()
    qs = (
        models.ItemCortesia.objects
        .values(produto=F('produtos__nome'))
        .annotate(
            total_qtd = Sum('quantidade')
        )
        .order_by('-quantidade')
    )

    if request.GET.get('export') == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Cortesias'

        ws.append(['Nome produto', 'Quantidade total'])

        for row in qs:
            ws.append([
                row['produto'],
                row['total_qtd']
            ])

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

        header_font = Font(bold=True, color='FFFFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='4F81BD')

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="cortesias.xlsx"'
        wb.save(response)
        return response
    table = tables.CortesiaTable(cortesias)
    RequestConfig(request, paginate={"per_page": 25}).configure(table)
    return render(request, 'cortesias/cortesias.html', {
        'title': 'Cortesias',
        'table': table
    })

@login_required
def add_cortesia(request):
    if request.method == 'POST':
        form = forms.CortesiaForm(request.POST)
        formset = ItemCortesiaCreatFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            cortesia = form.save(commit=False)
            cortesia.create_user = form.cleaned_data.get('create_user') or request.user
            cortesia.assign_user = form.cleaned_data.get('assign_user') or request.user
            cortesia.save()
            formset.instance = cortesia
            formset.save()
            messages.success(request, 'O novo registro de cortesia foi feito com sucesso!')
            return redirect('cortesias')
    else:
        form = forms.CortesiaForm()
        formset = ItemCortesiaCreatFormSet()
    return render(request, 'cortesias/add_cortesia.html', {
        'title': 'Adicionar cortesias',
        'form': form,
        'formset': formset
    })

@login_required
def edit_cortesia(request, cortesia_id):
    cortesia = get_object_or_404(models.Cortesia, pk=cortesia_id)
    if request.method == 'POST':
        form = forms.CortesiaForm(request.POST, instance=cortesia)
        formset = ItemCortesiaCreatFormSet(request.POST, instance=cortesia)
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                cortesia = form.save(commit=False)
                cortesia.assign_user = request.user
                cortesia.save()
                formset.save()
            messages.success(request, 'O registro de cortesia foi editado com sucesso!')
            return redirect('cortesias')
    else:
        form = forms.CortesiaForm(instance=cortesia)
        formset = ItemCortesiaCreatFormSet(instance=cortesia)
    return render(request, 'cortesias/add_cortesia.html', {
        'title': 'Editar registro de cortesia',
        'form': form,
        'formset': formset
    })

@login_required
def delete_cortesia_modal(request):
    pk = request.GET.get('id')
    record = get_object_or_404(models.Cortesia, id=pk)
    return render(request, 'cortesias/confirmar_deletar.html', {
        'record': record
    })

@login_required
def delete_cortesia(request, cortesia_id):
    qs_cortesia = get_object_or_404(models.Cortesia, pk=cortesia_id)
    if request.method == 'POST':
        qs_cortesia.delete()
        messages.success(request, 'O registro de cortesia foi deletado com sucesso!')
        return redirect('cortesias')
    return HttpResponse(status=405)

@login_required
def categoria_entrada(request):
    categoria_entrada = models.Categoria.objects.filter(is_entrada=True).order_by('nome').all()
    table = tables.CategoriaEntradaTable(categoria_entrada)
    RequestConfig(request, paginate={"per_page": 15}).configure(table)
    return render(request, 'categorias/categorias_entradas.html', {
        'title': 'Categorias',
        'table': table
    })

@login_required
def add_categoria_entrada(request):
    if request.method == 'POST':
        form = forms.CategoriaForm(request, request.POST)
        if form.is_valid():
            nova_categoria = form.save(commit=False)
            nova_categoria.is_despesa = False
            nova_categoria.is_entrada = True
            nova_categoria.create_user = form.cleaned_data.get('create_user') or request.user
            nova_categoria.assign_user = form.cleaned_data.get('assign_user') or request.user
            nova_categoria.save()
            messages.success(request, f'A categoria {nova_categoria.nome} foi adicionada com sucesso!')
            return redirect('categoria_entrada')
    else:
        form = forms.CategoriaForm(request)
    return render(request, 'categorias/add_categoria_entrada.html', {
        'title': 'Adicionar categoria',
        'form': form
    })

@login_required
def edit_categoria_entrada(request, categoria_entrada_id):
    qs_categoria = get_object_or_404(models.Categoria, pk=categoria_entrada_id)
    if request.method == 'POST':
        form = forms.CategoriaEditForm(request.POST, instance=qs_categoria)
        if form.is_valid():
            categoria = form.save(commit=False)
            categoria.assign_user = form.cleaned_data.get('assign_user') or request.user
            categoria.save()
            messages.success(request, f'A categoria {categoria.nome} foi editada com sucesso!')
            return redirect('categoria_entrada')
    else:
        form = forms.CategoriaEditForm(instance=qs_categoria)
    return render(request, 'categorias/add_categoria_entrada.html', {
        'title': 'Editar categoria',
        'form': form
    })

@login_required
def delete_categoria_entrada_modal(request):
    pk = request.GET.get('id')
    record = get_object_or_404(models.Categoria, id=pk)
    return render(request, 'categorias/confirmar_deletar_entrada.html', {
        'record': record
    })

@login_required
def delete_categoria_entrada(request, categoria_entrada_id):
    categoria = get_object_or_404(models.Categoria, pk=categoria_entrada_id)
    if request.method == 'POST':
        categoria.delete()
        messages.success(request, 'A categoria foi deletada com sucesso!')
        return redirect('categoria_entrada')
    return HttpResponse(status=405)

@login_required
def categoria_despesa(request):
    categoria_despesa = models.Categoria.objects.filter(is_despesa=True).order_by('nome').all()
    table = tables.CategoriaDespesaTable(categoria_despesa)
    RequestConfig(request, paginate={"per_page": 15}).configure(table)
    return render(request, 'categorias/categorias_despesas.html', {
        'title': 'Categorias',
        'table': table
    })

@login_required
def add_categoria_despesa(request):
    if request.method == 'POST':
        form = forms.CategoriaForm(request, request.POST)
        if form.is_valid():
            nova_categoria = form.save(commit=False)
            nova_categoria.is_despesa = True
            nova_categoria.is_entrada = False
            nova_categoria.create_user = form.cleaned_data.get('create_user') or request.user
            nova_categoria.assign_user = form.cleaned_data.get('assign_user') or request.user
            nova_categoria.save()
            messages.success(request, f'A categoria {nova_categoria.nome} foi adicionada com sucesso!')
            return redirect('categoria_despesa')
    else:
        form = forms.CategoriaForm(request)
    return render(request, 'categorias/add_categoria_despesa.html', {
        'title': 'Adicionar categoria',
        'form': form
    })

@login_required
def edit_categoria_despesa(request, categoria_despesa_id):
    qs_categoria = get_object_or_404(models.Categoria, pk=categoria_despesa_id)
    if request.method == 'POST':
        form = forms.CategoriaEditForm(request.POST, instance=qs_categoria)
        if form.is_valid():
            categoria = form.save(commit=False)
            categoria.assign_user = form.cleaned_data.get('assign_user') or request.user
            categoria.save()
            messages.success(request, f'A categoria {categoria.nome} foi editada com sucesso!')
            return redirect('categoria_despesa')
    else:
        form = forms.CategoriaEditForm(instance=qs_categoria)
    return render(request, 'categorias/add_categoria_despesa.html', {
        'title': 'Editar categoria',
        'form': form
    })

@login_required
def delete_categoria_despesa_modal(request):
    pk = request.GET.get('id')
    record = get_object_or_404(models.Categoria, id=pk)
    return render(request, 'categorias/confirmar_deletar_despesa.html', {
        'record': record
    })

@login_required
def delete_categoria_despesa(request, categoria_despesa_id):
    categoria = get_object_or_404(models.Categoria, pk=categoria_despesa_id)
    if request.method == 'POST':
        categoria.delete()
        messages.success(request, 'A categoria foi deletada com sucesso!')
        return redirect('categoria_despesa')
    return HttpResponse(status=405)

@login_required
def entradas(request):
    form = forms.EntradasFindForm(request.GET)
    form.fields['categoria'].required = False
    form.fields['data'].required = False
    filter_search = {}
    if form.is_valid():
        categoria = form.cleaned_data.get('categoria')
        data = form.cleaned_data.get('data')
        if categoria:
            filter_search['categoria'] = categoria
        if data:
            filter_search['data'] = data
    entradas = models.Entradas.objects.filter(**filter_search).order_by('categoria__nome').all()
    entradas_agrupadas = (
        entradas.values('categoria__nome')
        .annotate(total_entrada=Sum('valor'))
        .order_by('categoria__nome')
    )

    if request.GET.get('export') == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Entradas'

        ws.append(['Categoria', 'Total'])

        for row in entradas_agrupadas:
            ws.append([
                row['categoria__nome'],
                float(row['total_entrada'] or 0)
            ])

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

        row_count = len(entradas_agrupadas) + 1
        header_font = Font(bold=True, color='FFFFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='4F81BD')

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, max_row=row_count):
            for cell in row:
                cell.number_format = '"R$"#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')

        total_row = row_count + 2
        sum_range = f"B2:B{row_count}"
        cell_head = ws.cell(row=total_row, column=1, value="Total Geral:")
        cell_head.font = Font(bold=True)
        cell_head = Alignment(horizontal="right")
        cell = ws.cell(
            row=total_row,
            column=2,
            value=f"=SUM({sum_range})"
        )
        cell.number_format = '"R$"#,##0.00'
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right")

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="entradas.xlsx"'
        wb.save(response)
        return response
    soma_valor = entradas.aggregate(total_valor=Sum('valor'))['total_valor'] or Decimal('0.00')
    table = tables.EntradasTable(entradas)
    RequestConfig(request, paginate={"per_page": 25}).configure(table)
    return render(request, 'entradas/entradas.html', {
        'title': 'Entradas',
        'form': form,
        'table': table,
        'soma_valor': soma_valor
    })

@login_required
def add_entrada(request):
    if request.method == 'POST':
        form = forms.EntradasForm(request.POST)
        if form.is_valid():
            nova_entrada = form.save(commit=False)
            nova_entrada.create_user = form.cleaned_data.get('create_user') or request.user
            nova_entrada.assign_user = form.cleaned_data.get('assign_user') or request.user
            nova_entrada.save()
            messages.success(request, 'O novo registro de entradas foi feito com sucesso!')
            return redirect('entradas')
    else:
        form = forms.EntradasForm()
    return render(request, 'entradas/add_entrada.html', {
        'title': 'Adicionar entrada avulsa',
        'form': form
    })

@login_required
def edit_entrada(request, entrada_id):
    qs_entrada = get_object_or_404(models.Entradas, pk=entrada_id)
    if request.method == 'POST':
        form = forms.EntradasForm(request.POST, instance=qs_entrada)
        if form.is_valid():
            entrada = form.save(commit=False)
            entrada.assign_user = form.cleaned_data.get('assign_user') or request.user
            entrada.save()
            messages.success(request, 'O registro de entradas foi editado com sucesso!')
            return redirect('entradas')
    else:
        form = forms.EntradasForm(instance=qs_entrada)
    return render(request, 'entradas/add_entrada.html', {
        'title': 'Editar entrada avulsa',
        'form': form
    })

@login_required
def delete_entrada_modal(request):
    pk = request.GET.get('id')
    record = get_object_or_404(models.Entradas, id=pk)
    return render(request, 'entradas/confirmar_deletar.html', {
        'record': record
    })

@login_required
def delete_entrada(request, entrada_id):
    entrada = get_object_or_404(models.Entradas, pk=entrada_id)
    if request.method == 'POST':
        entrada.delete()
        messages.success(request, 'O registro da entrada foi deletado com sucesso!')
        return redirect('entradas')
    return HttpResponse(status=405)

@login_required
def despesas(request):
    form = forms.DespesasFindForm(request.GET)
    form.fields['categoria'].required = False
    form.fields['data'].required = False
    filter_search = {}
    if form.is_valid():
        categoria = form.cleaned_data.get('categoria')
        data = form.cleaned_data.get('data')
        is_pago = form.cleaned_data.get('is_pago')
        datapago = form.cleaned_data.get('datapago')
        if categoria:
            filter_search['categoria'] = categoria
        if data:
            filter_search['data'] = data
        if datapago:
            filter_search['datapago'] = datapago
        if is_pago:
            filter_search['is_pago'] = is_pago
    despesas = models.Despesas.objects.filter(**filter_search).order_by('categoria__nome').all()
    despesas_agrupadas = (
        despesas.values('categoria__nome')
        .annotate(total_entrada=Sum('valor'))
        .order_by('categoria__nome')
    )

    if request.GET.get('export') == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Despesas'

        ws.append(['Categoria', 'Total'])

        for row in despesas_agrupadas:
            ws.append([
                row['categoria__nome'],
                float(row['total_entrada'] or 0)
            ])

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

        row_count = len(despesas_agrupadas) + 1
        header_font = Font(bold=True, color='FFFFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='4F81BD')

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, max_row=row_count):
            for cell in row:
                cell.number_format = '"R$"#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')

        total_row = row_count + 2
        sum_range = f"B2:B{row_count}"
        cell_head = ws.cell(row=total_row, column=1, value="Total Geral:")
        cell_head.font = Font(bold=True)
        cell_head = Alignment(horizontal="right")
        cell = ws.cell(
            row=total_row,
            column=2,
            value=f"=SUM({sum_range})"
        )
        cell.number_format = '"R$"#,##0.00'
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right")

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="despesas.xlsx"'
        wb.save(response)
        return response
    soma_valor = despesas.aggregate(total_valor=Sum('valor'))['total_valor'] or Decimal('0.00')
    table = tables.DespesasTable(despesas)
    RequestConfig(request, paginate={"per_page": 15}).configure(table)
    return render(request, 'despesas/despesas.html', {
        'title': 'Despesas',
        'soma_valor': soma_valor,
        'table': table,
        'form': form
    })

@login_required
def add_despesa(request):
    if request.method == 'POST':
        form = forms.DespesasForm(request.POST, request.FILES)
        if form.is_valid():
            nova_despesa = form.save(commit=False)
            nova_despesa.create_user = form.cleaned_data.get('create_user') or request.user
            nova_despesa.assign_user = form.cleaned_data.get('assign_user') or request.user
            nova_despesa.save()
            messages.success(request, 'O novo registro de despesas foi feito com sucesso!')
            return redirect('despesas')
    else:
        form = forms.DespesasForm()
    return render(request, 'despesas/add_despesa.html', {
        'title': 'Adicionar despesa',
        'form': form
    })

@login_required
def edit_despesa(request, despesa_id):
    qs_despesa = get_object_or_404(models.Despesas, pk=despesa_id)
    if request.method == 'POST':
        form = forms.DespesasForm(request.POST, request.FILES, instance=qs_despesa)
        if form.is_valid():
            despesa = form.save(commit=False)
            despesa.assign_user = form.cleaned_data.get('assign_user') or request.user
            despesa.save()
            messages.success(request, 'O registro da despesa foi editado com sucesso!')
            return redirect('despesas')
    else:
        form = forms.DespesasForm(instance=qs_despesa)
    return render(request, 'despesas/add_despesa.html', {
        'title': 'Editar despesa',
        'form': form
    })

@login_required
def delete_despesa_modal(request):
    pk = request.GET.get('id')
    record = get_object_or_404(models.Despesas, id=pk)
    return render(request, 'despesas/confirmar_deletar.html', {
        'record': record
    })

@login_required
def delete_despesa(request, despesa_id):
    despesa = get_object_or_404(models.Despesas, pk=despesa_id)
    if request.method == 'POST':
        despesa.delete()
        messages.success(request, 'O registro de despesa foi excluido com sucesso')
        return redirect('despesas')
    return HttpResponse(status=405)